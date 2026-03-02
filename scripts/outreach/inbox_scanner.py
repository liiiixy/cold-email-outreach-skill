"""
收件箱扫描器 — 发送后自动检测退信和自动回复。

用法:
    from inbox_scanner import scan_inbox, classify_responses, generate_tracking_report

流程:
    1. scan_inbox() — IMAP 连接并拉取最近的邮件
    2. classify_responses() — 将邮件分为退信/自动回复/无关
    3. generate_tracking_report() — 合并发送列表 + 收件结果 → 追踪报告
"""

import imaplib
import email as email_lib
from email.header import decode_header
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import re
import json


def scan_inbox(
    imap_host: str,
    username: str,
    password: str,
    since_hours: int = 24,
    imap_port: int = 993,
) -> List[Dict]:
    """扫描收件箱，返回最近的邮件列表。"""
    imap = imaplib.IMAP4_SSL(imap_host, imap_port)
    imap.login(username, password)
    imap.select("INBOX")

    since_date = (datetime.now() - timedelta(hours=since_hours)).strftime("%d-%b-%Y")
    _, msg_ids = imap.search(None, f'(SINCE "{since_date}")')

    results = []
    for eid in msg_ids[0].split():
        _, data = imap.fetch(eid, "(RFC822)")
        msg = email_lib.message_from_bytes(data[0][1])

        subject = _decode_header(msg["Subject"] or "")
        from_addr = msg["From"] or ""
        to_addr = msg["To"] or ""
        date_str = msg["Date"] or ""

        body = _get_body(msg)

        results.append({
            "subject": subject,
            "from": from_addr,
            "to": to_addr,
            "date": date_str,
            "body_preview": body[:500],
        })

    imap.logout()
    return results


def classify_responses(
    inbox_emails: List[Dict],
    sent_emails: List[Dict],
    retried_emails: Optional[Dict[str, str]] = None,
) -> Dict:
    """
    将收件箱邮件与发送列表匹配，分为三类:
    - bounced: 发送失败（退信且未修复）
    - auto_replied: 已送达（收到自动回复）
    - no_response: 暂无回复

    已修复的退信会被静默忽略，新邮箱按正常流程归类。

    参数:
        inbox_emails: scan_inbox() 的返回结果
        sent_emails: /tmp/outreach_send_list.json 的内容
        retried_emails: 已修复的退信映射 {旧邮箱: 新邮箱}，
                        如果为 None，会自动从 /tmp/outreach_retried.json 读取
    """
    # 加载已修复的退信记录
    if retried_emails is None:
        retried_emails = _load_retried_emails()

    # 建立发送列表的邮箱 → 品牌映射
    sent_map = {}
    for item in sent_emails:
        m = item["merchant"]
        em = m.get("email", "")
        sent_map[em.lower()] = m.get("brand_name", "")

    # 所有发送过的域名
    sent_domains = {em.split("@")[1].lower() for em in sent_map if "@" in em}

    # 已修复退信的旧邮箱集合（静默忽略这些退信）
    retried_old_emails = {k.lower() for k in retried_emails}

    bounced = []
    auto_replied = []
    matched_emails = set()

    for msg in inbox_emails:
        from_addr = msg["from"].lower()
        subject = msg["subject"]

        # ── 退信检测 ──
        if _is_bounce(from_addr, subject):
            failed_email = _extract_bounced_email(msg["body_preview"])
            if not failed_email:
                continue
            # 已修复的退信 → 静默跳过
            if failed_email.lower() in retried_old_emails:
                continue
            brand = sent_map.get(failed_email.lower(), "Unknown")
            bounced.append({
                "brand": brand,
                "email": failed_email,
                "reason": _extract_bounce_reason(msg["body_preview"]),
            })
            matched_emails.add(failed_email.lower())
            continue

        # ── 自动回复检测 ──
        from_domain = _extract_domain(from_addr)
        if from_domain and from_domain in sent_domains:
            brand = _find_brand_by_domain(from_domain, sent_map)
            if brand:
                email_key = _find_email_by_domain(from_domain, sent_map)
                auto_replied.append({
                    "brand": brand,
                    "email": email_key,
                    "response_summary": subject[:100],
                })
                if email_key:
                    matched_emails.add(email_key.lower())

    # ── 暂无回复 ──
    no_response = []
    for em, brand in sent_map.items():
        if em.lower() not in matched_emails:
            no_response.append({"brand": brand, "email": em})

    total_sent = len(sent_emails)

    return {
        "bounced": bounced,
        "auto_replied": auto_replied,
        "no_response": no_response,
        "total_sent": total_sent,
        "delivered": total_sent - len(bounced),
        "delivery_rate": (total_sent - len(bounced)) * 100 // max(total_sent, 1),
    }


def generate_tracking_report(classification: Dict) -> str:
    """生成终端可读的追踪报告。"""
    c = classification
    lines = []
    lines.append("=" * 60)
    lines.append("📊 OUTREACH 追踪报告")
    lines.append("=" * 60)
    lines.append(f"总计发送: {c['total_sent']} 封")
    lines.append(f"✅ 已送达: {len(c['auto_replied'])}")
    if c["bounced"]:
        lines.append(f"❌ 失败: {len(c['bounced'])}")
    lines.append(f"⏳ 暂无回复: {len(c['no_response'])}")
    lines.append(f"📬 送达率: {c['delivery_rate']}%")
    lines.append("=" * 60)

    if c["auto_replied"]:
        lines.append(f"\n✅ 已送达 ({len(c['auto_replied'])})")
        lines.append("-" * 60)
        for ar in c["auto_replied"]:
            lines.append(f"  {ar['brand']:<30} {ar['email']}")

    if c["bounced"]:
        lines.append(f"\n❌ 失败 ({len(c['bounced'])})")
        lines.append("-" * 60)
        for b in c["bounced"]:
            lines.append(f"  {b['brand']:<30} {b['email']}")
            lines.append(f"    原因: {b['reason']}")

    if c["no_response"]:
        lines.append(f"\n⏳ 暂无回复 ({len(c['no_response'])})")
        lines.append("-" * 60)
        for nr in c["no_response"]:
            lines.append(f"  {nr['brand']:<30} {nr['email']}")

    lines.append("\n" + "=" * 60)
    return "\n".join(lines)


def export_tracking_excel(
    classification: Dict,
    send_list: List[Dict],
    output_path: str,
):
    """导出追踪报告到 Excel。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: 追踪状态 ──
    ws1 = wb.active
    ws1.title = "追踪状态"

    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["#", "Brand", "Email", "Status", "Detail"]
    widths = [4, 28, 38, 16, 50]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill = hfill
        cell.font = hfont
        ws1.column_dimensions[get_column_letter(ci)].width = w

    c = classification
    row = 2
    idx = 1

    for ar in c["auto_replied"]:
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=ar["brand"])
        ws1.cell(row=row, column=3, value=ar["email"])
        cell = ws1.cell(row=row, column=4, value="✅ 已送达")
        cell.fill = green
        ws1.cell(row=row, column=5, value="").alignment = wrap
        row += 1; idx += 1

    for b in c["bounced"]:
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=b["brand"])
        ws1.cell(row=row, column=3, value=b["email"])
        cell = ws1.cell(row=row, column=4, value="❌ 失败")
        cell.fill = red
        ws1.cell(row=row, column=5, value=b["reason"]).alignment = wrap
        row += 1; idx += 1

    for nr in c["no_response"]:
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=nr["brand"])
        ws1.cell(row=row, column=3, value=nr["email"])
        cell = ws1.cell(row=row, column=4, value="⏳ 暂无回复")
        cell.fill = yellow
        ws1.cell(row=row, column=5, value="")
        row += 1; idx += 1

    ws1.auto_filter.ref = f"A1:E{row - 1}"

    # ── Sheet 2: 邮件内容 ──
    ws2 = wb.create_sheet("邮件内容")
    h2 = ["#", "Brand", "Email", "Subject", "Body", "Words"]
    w2 = [4, 24, 34, 48, 90, 6]
    for ci, (h, w) in enumerate(zip(h2, w2), 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = hfill; cell.font = hfont
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ri, item in enumerate(send_list, 2):
        m = item["merchant"]; ec = item["email_content"]
        vals = [ri-1, m.get("brand_name",""), m.get("email",""),
                ec["subject"], ec["body"], ec.get("word_count","")]
        for ci, v in enumerate(vals, 1):
            ws2.cell(row=ri, column=ci, value=v).alignment = wrap
        ws2.row_dimensions[ri].height = 100

    # ── Sheet 3: 发送摘要 ──
    ws3 = wb.create_sheet("发送摘要")
    stats = [
        ("总计发送", c["total_sent"]),
        ("已送达", len(c["auto_replied"])),
        ("失败", len(c["bounced"])),
        ("暂无回复", len(c["no_response"])),
        ("送达率", f"{c['delivery_rate']}%"),
    ]
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15
    for ri, (label, val) in enumerate(stats, 1):
        ws3.cell(row=ri, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=val)

    wb.save(output_path)
    return output_path


RETRIED_FILE = "/tmp/outreach_retried.json"


def save_retried_emails(retried_map: Dict[str, str]):
    """保存已修复退信的映射到文件。每次重发成功后调用。

    参数:
        retried_map: {旧邮箱: 新邮箱}，如 {"nshop@fearofgod.com": "info@fearofgod.com"}
    """
    existing = _load_retried_emails()
    existing.update({k.lower(): v for k, v in retried_map.items()})
    with open(RETRIED_FILE, "w") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)


def _load_retried_emails() -> Dict[str, str]:
    """从文件加载已修复退信的映射。"""
    try:
        with open(RETRIED_FILE) as f:
            data = json.load(f)
            return {k.lower(): v for k, v in data.items()}
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


# ── Private helpers ─────────────────────────────────────────────────

def _decode_header(raw: str) -> str:
    decoded = decode_header(raw)
    result = ""
    for part, enc in decoded:
        if isinstance(part, bytes):
            result += part.decode(enc or "utf-8", errors="replace")
        else:
            result += part
    return result


def _get_body(msg) -> str:
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    return payload.decode("utf-8", errors="replace")
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            return payload.decode("utf-8", errors="replace")
    return ""


def _is_bounce(from_addr: str, subject: str) -> bool:
    from_lower = from_addr.lower()
    subject_lower = subject.lower()
    return (
        "mailer-daemon" in from_lower
        or "postmaster" in from_lower
        or "delivery status notification" in subject_lower
        or "undeliverable" in subject_lower
        or "mail delivery failed" in subject_lower
    )


def _extract_bounced_email(body: str) -> str:
    """从退信正文中提取失败的目标邮箱。"""
    # Pattern: "wasn't delivered to xxx@yyy.com"
    m = re.search(r"wasn't delivered to\s+(\S+@\S+)", body)
    if m:
        return m.group(1).strip().rstrip(".")
    # Pattern: "could not be delivered to: xxx@yyy.com"
    m = re.search(r"delivered to:?\s+(\S+@\S+)", body)
    if m:
        return m.group(1).strip().rstrip(".")
    # Generic email pattern after "failed" or "bounce"
    m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", body)
    if m:
        return m.group(0)
    return ""


def _extract_bounce_reason(body: str) -> str:
    if "address couldn't be found" in body.lower() or "address not found" in body.lower():
        return "Address not found"
    if "mailbox full" in body.lower():
        return "Mailbox full"
    if "rejected" in body.lower():
        return "Rejected by server"
    return "Delivery failed"


def _extract_domain(from_str: str) -> Optional[str]:
    m = re.search(r"@([\w.-]+)", from_str)
    if m:
        domain = m.group(1).lower()
        # Skip common email service domains
        skip = {"googlemail.com", "gmail.com", "outlook.com", "yahoo.com",
                "zendesk.com", "reamaze.com", "freshdesk.com", "helpscout.net"}
        # For zendesk etc, extract subdomain: support@moussy-global.zendesk.com → moussy-global
        for s in skip:
            if domain.endswith(s) and domain != s:
                subdomain = domain.replace(f".{s}", "")
                return subdomain
        if domain not in skip:
            return domain
    return None


def _find_brand_by_domain(domain: str, sent_map: dict) -> Optional[str]:
    for em, brand in sent_map.items():
        em_domain = em.split("@")[1].lower() if "@" in em else ""
        if em_domain == domain or domain in em_domain or em_domain in domain:
            return brand
    return None


def _find_email_by_domain(domain: str, sent_map: dict) -> Optional[str]:
    for em in sent_map:
        em_domain = em.split("@")[1].lower() if "@" in em else ""
        if em_domain == domain or domain in em_domain or em_domain in domain:
            return em
    return None


# ── V2: 智能分类（区分自动回复 vs 真人回复）──────────────────────────

AUTO_REPLY_SIGNALS = [
    r"we have received your (email|inquiry|message)",
    r"this is an auto(matic|mated) (response|reply|message)",
    r"will (respond|reply|get back to you) within \d+",
    r"(24|48|72)[\s-]*(hours?|business days?)",
    r"ticket ?(id|#|number)?[:\s]*[#]?\d+",
    r"\[[\w]+-[\w]+\]",  # ticket IDs like [KJNK3M-P77X5]
    r"our (customer (care|service)|support) (team|hours|representatives?)",
    r"(monday|tuesday|wednesday|thursday|friday).*(am|pm)",
    r"check our faq",
    r"(shipping|return|exchange) (policy|page|site)",
]

HUMAN_REPLY_SIGNALS = [
    r"forwarded (this|your) (email|inquiry|message) to",
    r"(relevant|appropriate) department",
    r"contact (our|the) .*(team|department) (directly|at)",
    r"you can (find|reach) them at",
    r"mailto:",
    r"(tell|share|show) (me|us) more",
    r"(set up|schedule|arrange) a (demo|call|meeting)",
    r"not interested",
    r"(best wishes|kind regards|warm regards),?\s*\n\s*\w+",  # personal sign-off with name
]


def _is_auto_reply(body: str, subject: str, time_delta_seconds: float) -> bool:
    """判断是否为自动回复。"""
    body_lower = body.lower()
    subject_lower = subject.lower()
    text = body_lower + " " + subject_lower

    score = 0
    for pattern in AUTO_REPLY_SIGNALS:
        if re.search(pattern, text):
            score += 1
    # 发送后 5 分钟内回复 → 大概率自动回复
    if 0 < time_delta_seconds < 300:
        score += 1
    return score >= 2


def _is_human_reply(body: str, subject: str, time_delta_seconds: float) -> bool:
    """判断是否为真人回复。"""
    body_lower = body.lower()
    subject_lower = subject.lower()
    text = body_lower + " " + subject_lower

    for pattern in HUMAN_REPLY_SIGNALS:
        if re.search(pattern, text):
            return True
    # 发送后 1 小时以上回复 + 不像自动回复 → 可能是真人
    if time_delta_seconds > 3600 and not _is_auto_reply(body, subject, time_delta_seconds):
        return True
    return False


def _detect_intent(body: str, subject: str) -> str:
    """检测回复意图。"""
    text = (body + " " + subject).lower()

    if re.search(r"forwarded|relevant department|appropriate team", text):
        return "forwarded_internally"
    if re.search(r"contact .*(team|department) (directly|at)|mailto:|@[\w.-]+\.com", text):
        # 检查是否包含新的邮箱地址（非发件人自己的）
        emails = re.findall(r"[\w.+-]+@[\w-]+\.[\w.-]+", text)
        if emails:
            return "referral"
    if re.search(r"(tell|share|show) (me|us) more|interested|demo|schedule|call", text):
        return "interested"
    if re.search(r"not interested|not .* right time|no thank|unsubscribe|remove me", text):
        return "declined"
    return "acknowledged"


def _extract_referral_email(body: str) -> Optional[str]:
    """从回复中提取转介绍的新邮箱。"""
    # 寻找 mailto: 链接
    m = re.search(r"mailto:([\w.+-]+@[\w-]+\.[\w.-]+)", body.lower())
    if m:
        return m.group(1)
    # 寻找文本中的邮箱（排除常见的通用地址模式）
    emails = re.findall(r"[\w.+-]+@[\w-]+\.[\w.-]+", body)
    for em in emails:
        em_lower = em.lower()
        if not any(x in em_lower for x in ["noreply", "no-reply", "mailer-daemon",
                                            "support@", "help@", "info@", "customercare@"]):
            return em
    return None


def _match_reply_to_sent(from_addr: str, sent_map: dict) -> Optional[tuple]:
    """
    将收件箱回复匹配到发送列表。返回 (brand, email_key) 或 None。

    匹配策略（按优先级）：
    1. 精确邮箱匹配（如 reply from passionliliemarketing@gmail.com）
    2. 域名精确匹配（如 reply from contact@toa.st → sent to contact@toa.st）
    3. 模糊域名匹配（如 reply from support@toddsnyder.reamaze.com → sent to info@toddsnyder.com）
    """
    # 提取回复邮箱
    reply_email = re.search(r"[\w.+-]+@[\w.-]+", from_addr)
    if not reply_email:
        return None
    reply_email = reply_email.group(0).lower()

    # 策略 1: 精确邮箱匹配
    if reply_email in sent_map:
        return (sent_map[reply_email], reply_email)

    # 提取回复域名
    reply_domain = reply_email.split("@")[1] if "@" in reply_email else ""

    # 策略 2: 域名精确匹配
    for em, brand in sent_map.items():
        em_domain = em.split("@")[1] if "@" in em else ""
        if em_domain == reply_domain:
            return (brand, em)

    # 策略 3: 模糊匹配 — 处理 reamaze/zendesk/freshdesk 等代理域名
    hosted_platforms = {"zendesk.com", "reamaze.com", "freshdesk.com",
                        "helpscout.net", "intercom.io", "kayako.com"}
    for platform in hosted_platforms:
        if reply_domain.endswith(platform):
            # 提取子域名: toddsnyder.reamaze.com → toddsnyder
            subdomain = reply_domain.replace(f".{platform}", "")
            for em, brand in sent_map.items():
                em_domain = em.split("@")[1] if "@" in em else ""
                # toddsnyder 在 toddsnyder.com 中
                em_base = em_domain.split(".")[0] if em_domain else ""
                if subdomain == em_base or subdomain in em_domain or em_base in subdomain:
                    return (brand, em)

    # 策略 4: 域名基础名模糊匹配（如 shopmashburn.com vs sidmashburn.com）
    reply_base = reply_domain.split(".")[0] if reply_domain else ""
    for em, brand in sent_map.items():
        em_domain = em.split("@")[1] if "@" in em else ""
        em_base = em_domain.split(".")[0] if em_domain else ""
        # 至少 5 字符的共同子串匹配
        if len(reply_base) >= 5 and len(em_base) >= 5:
            if reply_base in em_base or em_base in reply_base:
                return (brand, em)

    return None


def classify_responses_v2(
    inbox_emails: List[Dict],
    sent_emails: List[Dict],
    retried_emails: Optional[Dict[str, str]] = None,
) -> Dict:
    """
    V2 分类：区分自动回复和真人回复，并检测回复意图。

    返回 4 类：bounced / auto_reply / human_reply / no_response
    """
    if retried_emails is None:
        retried_emails = _load_retried_emails()

    sent_map = {}
    for item in sent_emails:
        m = item["merchant"]
        em = m.get("email", "")
        sent_map[em.lower()] = m.get("brand_name", "")

    retried_old_emails = {k.lower() for k in retried_emails}

    bounced = []
    auto_reply = []
    human_reply = []
    matched_emails = set()

    for msg in inbox_emails:
        from_addr = msg["from"].lower()
        subject = msg["subject"]
        body = msg["body_preview"]

        # ── 退信检测 ──
        if _is_bounce(from_addr, subject):
            failed_email = _extract_bounced_email(body)
            if not failed_email:
                continue
            if failed_email.lower() in retried_old_emails:
                continue
            brand = sent_map.get(failed_email.lower(), "Unknown")
            bounced.append({
                "brand": brand,
                "email": failed_email,
                "reason": _extract_bounce_reason(body),
            })
            matched_emails.add(failed_email.lower())
            continue

        # ── 匹配发送列表（使用改进的匹配逻辑）──
        match = _match_reply_to_sent(from_addr, sent_map)
        if not match:
            continue

        brand, email_key = match
        time_delta = 3600  # 默认值

        intent = _detect_intent(body, subject)
        referral_email = None
        if intent == "referral":
            referral_email = _extract_referral_email(body)

        entry = {
            "brand": brand,
            "email": email_key,
            "subject": subject[:100],
            "body_preview": body[:300],
            "intent": intent,
        }
        if referral_email:
            entry["referral_email"] = referral_email

        if _is_human_reply(body, subject, time_delta):
            human_reply.append(entry)
        else:
            auto_reply.append(entry)

        if email_key:
            matched_emails.add(email_key.lower())

    # ── 暂无回复 ──
    no_response = []
    for em, brand in sent_map.items():
        if em.lower() not in matched_emails:
            no_response.append({"brand": brand, "email": em})

    total_sent = len(sent_emails)

    return {
        "bounced": bounced,
        "auto_reply": auto_reply,
        "human_reply": human_reply,
        "no_response": no_response,
        "total_sent": total_sent,
        "delivered": total_sent - len(bounced),
        "delivery_rate": (total_sent - len(bounced)) * 100 // max(total_sent, 1),
    }


def generate_tracking_report_v2(classification: Dict) -> str:
    """生成 V2 追踪报告（区分自动回复和真人回复）。"""
    c = classification
    lines = []
    lines.append("=" * 60)
    lines.append("📊 OUTREACH 追踪报告")
    lines.append("=" * 60)
    lines.append(f"总计发送: {c['total_sent']} 封 | 送达率: {c['delivery_rate']}%")
    lines.append("")

    if c["human_reply"]:
        lines.append(f"⭐ 真人回复 ({len(c['human_reply'])}) — 需要跟进")
        lines.append("-" * 60)
        for hr in c["human_reply"]:
            intent_label = {
                "forwarded_internally": "已转内部",
                "referral": "转介绍",
                "interested": "感兴趣",
                "declined": "拒绝",
                "acknowledged": "已确认",
            }.get(hr["intent"], hr["intent"])
            line = f"  {hr['brand']:<28} {hr['email']:<32} {intent_label}"
            if hr.get("referral_email"):
                line += f" → {hr['referral_email']}"
            lines.append(line)
        lines.append("")

    if c["auto_reply"]:
        lines.append(f"📬 自动回复 ({len(c['auto_reply'])}) — 等待人工回复")
        lines.append("-" * 60)
        for ar in c["auto_reply"]:
            lines.append(f"  {ar['brand']:<28} {ar['email']:<32} {ar['subject'][:40]}")
        lines.append("")

    if c["bounced"]:
        lines.append(f"❌ 退信 ({len(c['bounced'])})")
        lines.append("-" * 60)
        for b in c["bounced"]:
            lines.append(f"  {b['brand']:<28} {b['email']:<32} {b['reason']}")
        lines.append("")

    if c["no_response"]:
        lines.append(f"⏳ 暂无回复 ({len(c['no_response'])})")
        lines.append("-" * 60)
        for nr in c["no_response"]:
            lines.append(f"  {nr['brand']:<28} {nr['email']}")

    lines.append("\n" + "=" * 60)
    return "\n".join(lines)
