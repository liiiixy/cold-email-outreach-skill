"""
reporter.py - 发送报告生成

输出：成功/失败/bounce/未获取邮箱的完整统计。
支持导出为 Excel 或格式化文本。
"""

import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_report(send_results: list[dict], merchants: list[dict] = None) -> dict:
    """
    生成发送报告统计。

    参数:
        send_results: batch_send 返回的结果列表
        merchants: 完整商家列表（用于统计未获取邮箱的）

    返回: 报告字典
    """
    total_attempted = len(send_results)
    successful = [r for r in send_results if r.get("success")]
    failed = [r for r in send_results if not r.get("success") and r.get("email")]
    no_email = [r for r in send_results if not r.get("email")]

    # 错误分类
    error_types = {}
    for r in failed:
        err = r.get("error", "unknown")
        # 简单分类
        if "bounce" in err.lower() or "550" in err or "undeliverable" in err.lower():
            category = "bounce"
        elif "rate" in err.lower() or "limit" in err.lower() or "429" in err:
            category = "rate_limited"
        elif "auth" in err.lower() or "401" in err or "403" in err:
            category = "auth_error"
        elif "timeout" in err.lower():
            category = "timeout"
        else:
            category = "other"
        error_types[category] = error_types.get(category, 0) + 1

    # 未获取邮箱的商家
    no_email_merchants = []
    if merchants:
        for m in merchants:
            if not m.get("email"):
                no_email_merchants.append({
                    "brand_name": m.get("brand_name", "未知"),
                    "lookup_status": m.get("email_lookup_status", "unknown"),
                })

    return {
        "timestamp": datetime.now().isoformat(),
        "summary": {
            "total_merchants": len(merchants) if merchants else total_attempted,
            "total_attempted": total_attempted,
            "successful": len(successful),
            "failed": len(failed),
            "no_email": len(no_email),
            "success_rate": f"{len(successful)/total_attempted*100:.1f}%" if total_attempted > 0 else "0%",
        },
        "error_breakdown": error_types,
        "successful_sends": [
            {"merchant": r["merchant_name"], "email": r["email"], "message_id": r.get("message_id", "")}
            for r in successful
        ],
        "failed_sends": [
            {"merchant": r["merchant_name"], "email": r["email"], "error": r.get("error", "")}
            for r in failed
        ],
        "no_email_merchants": no_email_merchants,
    }


def format_report_text(report: dict) -> str:
    """将报告格式化为文本。"""
    s = report["summary"]
    lines = []

    lines.append("=" * 50)
    lines.append("📊 邮件发送报告")
    lines.append("=" * 50)
    lines.append(f"⏰ 时间: {report['timestamp'][:19]}")
    lines.append("")
    lines.append("## 总体统计")
    lines.append(f"  商家总数:   {s['total_merchants']}")
    lines.append(f"  尝试发送:   {s['total_attempted']}")
    lines.append(f"  ✅ 成功:    {s['successful']}")
    lines.append(f"  ❌ 失败:    {s['failed']}")
    lines.append(f"  ⚠️ 无邮箱:  {s['no_email']}")
    lines.append(f"  📈 成功率:  {s['success_rate']}")

    if report["error_breakdown"]:
        lines.append("")
        lines.append("## 失败原因分析")
        error_labels = {
            "bounce": "退信 (Bounce)",
            "rate_limited": "限流 (Rate Limited)",
            "auth_error": "认证错误",
            "timeout": "超时",
            "other": "其他",
        }
        for err_type, count in report["error_breakdown"].items():
            label = error_labels.get(err_type, err_type)
            lines.append(f"  - {label}: {count} 封")

    if report["successful_sends"]:
        lines.append("")
        lines.append("## 发送成功列表")
        for item in report["successful_sends"]:
            lines.append(f"  ✅ {item['merchant']} → {item['email']}")

    if report["failed_sends"]:
        lines.append("")
        lines.append("## 发送失败列表")
        for item in report["failed_sends"]:
            lines.append(f"  ❌ {item['merchant']} → {item['email']}")
            lines.append(f"     原因: {item['error'][:80]}")

    if report["no_email_merchants"]:
        lines.append("")
        lines.append("## 未获取邮箱的商家")
        for item in report["no_email_merchants"]:
            status_map = {
                "not_found": "网页未找到",
                "no_source": "无官网/contact页面",
                "fetch_failed": "页面访问失败",
            }
            status = status_map.get(item["lookup_status"], item["lookup_status"])
            lines.append(f"  ⚠️ {item['brand_name']} ({status})")

    lines.append("")
    lines.append("=" * 50)
    return "\n".join(lines)


def export_to_excel(report: dict, send_list: list[dict], output_path: str) -> str:
    """
    将报告导出为 Excel 文件。

    参数:
        report: generate_report 返回的报告
        send_list: 完整的发送列表（含商家信息和邮件内容）
        output_path: 输出文件路径

    返回: 输出文件路径
    """
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl: pip3 install openpyxl")

    wb = openpyxl.Workbook()

    # ── Sheet 1: 总结 ──
    ws_summary = wb.active
    ws_summary.title = "发送总结"

    # 样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    success_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", fill_type="solid")

    s = report["summary"]
    summary_data = [
        ("指标", "数值"),
        ("商家总数", s["total_merchants"]),
        ("尝试发送", s["total_attempted"]),
        ("发送成功", s["successful"]),
        ("发送失败", s["failed"]),
        ("无邮箱", s["no_email"]),
        ("成功率", s["success_rate"]),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws_summary.cell(row=row_idx, column=1).font = header_font
            ws_summary.cell(row=row_idx, column=2).font = header_font

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20

    # ── Sheet 2: 详细记录 ──
    ws_detail = wb.create_sheet("发送详情")
    headers = ["商家名", "邮箱", "邮件标题", "发送状态", "错误信息", "品类", "优先级"]

    for col, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font_white

    row = 2
    for item in send_list:
        merchant = item.get("merchant", {})
        content = item.get("email_content", {})
        result = item.get("send_result", {})

        name = merchant.get("brand_name", "")
        email = merchant.get("email", "")
        subject = content.get("subject", "")
        success = result.get("success", False)
        error = result.get("error", "")
        cats = merchant.get("categories", "")
        priority = merchant.get("priority", "")

        ws_detail.cell(row=row, column=1, value=name)
        ws_detail.cell(row=row, column=2, value=email)
        ws_detail.cell(row=row, column=3, value=subject)

        status_cell = ws_detail.cell(row=row, column=4, value="成功" if success else ("无邮箱" if not email else "失败"))
        if success:
            status_cell.fill = success_fill
        elif not email:
            status_cell.fill = warn_fill
        else:
            status_cell.fill = fail_fill

        ws_detail.cell(row=row, column=5, value=error)
        ws_detail.cell(row=row, column=6, value=cats)
        ws_detail.cell(row=row, column=7, value=priority)
        row += 1

    # 自动列宽
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_detail.column_dimensions[col_letter].width = 25

    # ── 保存 ──
    wb.save(output_path)
    return output_path


# CLI 入口
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法:")
        print("  python3 reporter.py --text <报告JSON>")
        print("  python3 reporter.py --excel <报告JSON> <发送列表JSON> <输出路径>")
        sys.exit(1)

    if sys.argv[1] == "--text":
        with open(sys.argv[2]) as f:
            report = json.load(f)
        print(format_report_text(report))
    elif sys.argv[1] == "--excel":
        with open(sys.argv[2]) as f:
            report = json.load(f)
        with open(sys.argv[3]) as f:
            send_list = json.load(f)
        output = sys.argv[4] if len(sys.argv) > 4 else "outreach_report.xlsx"
        path = export_to_excel(report, send_list, output)
        print(f"报告已导出: {path}")
